import streamlit as st
import pandas as pd
from datetime import datetime
import os
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# =========================================================
# CẤU HÌNH TRANG WEB
# =========================================================
st.set_page_config(page_title="Quản Lý Tàu Cá NH-VN", page_icon="🚢", layout="wide")

# =========================================================
# THIẾT LẬP CƠ SỞ DỮ LIỆU LƯU TRỮ TỰ ĐỘNG
# =========================================================
APP_DIR = os.path.join(os.path.expanduser('~'), "PhanMem_TauCa_NHVN")
os.makedirs(APP_DIR, exist_ok=True)
DB_FILE = os.path.join(APP_DIR, "CSDL_TauCa_Local.xlsx")

# =========================================================
# 1. BỘ TỪ ĐIỂN TÌM KIẾM AI
# =========================================================
COLUMN_ALIASES = {
    'SO_DANG_KY': ['số đăng ký', 'biển số', 'số đk'],
    'CHU_TAU': ['chủ tàu', 'chủ phương tiện', 'họ tên chủ', 'tên chủ'],
    'SDT': ['số điện thoại', 'sđt', 'sdt', 'điện thoại'],
    'CCCD': ['cccd', 'cmnd', 'căn cước', 'chứng minh', 'số định danh', 'số cmnd'],
    'DIA_CHI': ['địa chỉ', 'nơi thường trú', 'địa chỉ thường trú', 'chỗ ở', 'địa chỉ chủ'],
    'LMAX': ['chiều dài lmax', 'lmax', 'chiều dài lớn nhất', 'l max', 'chiều dài'],
    'CONG_SUAT': ['công suất', 'cv', 'kw', 'máy chính'],
    'HAN_DK': ['hạn đăng kiểm', 'hết hạn đk', 'ngày hết hạn đăng kiểm (dd/mm/yyyy)', 'ngày hết hạn đăng kiểm', 'hạn đk'],
    'HAN_GP': ['hạn giấy phép', 'hạn gp', 'hết hạn gp', 'hạn gpktts', 'giấy phép ktts', 'ngày hết hạn (đối chiếu)', 'ngày hết hạn'],
    'NGAY_HET_HAN': ['hết hạn']
}

GLOBAL_EXCLUDES = ["phú yên", "ninh thuận", "bình thuận", "đắk lắk", "lâm đồng", "tp.hcm", "hồ chí minh", "hà nội", "đà nẵng", "quảng nam", "quảng ngãi", "bình định"]
KH_EXCLUDES = ["nha trang", "cam ranh", "diên khánh", "cam lâm", "khánh vĩnh", "khánh sơn", "xã ninh hải"]
mapping_rules = {
    "Xã Đại Lãnh": {"keywords": ["vạn thạnh", "vạn thọ", "đại lãnh"], "exclude": KH_EXCLUDES},
    "Xã Tu Bông": {"keywords": ["vạn khánh", "vạn long", "vạn phước", "tu bông"], "exclude": KH_EXCLUDES},
    "Xã Vạn Hưng": {"keywords": ["xuân sơn", "vạn hưng"], "exclude": KH_EXCLUDES},
    "Xã Vạn Ninh": {"keywords": ["vạn giã", "tt vạn giã", "thị trấn vạn giã", "vạn phú", "vạn lương", "xã vạn ninh"], "exclude": KH_EXCLUDES},
    "Xã Vạn Thắng": {"keywords": ["vạn bình", "vạn thắng"], "exclude": KH_EXCLUDES},
    "Phường Đông Ninh Hoà": {"keywords": ["ninh diêm", "ninh hải", "ninh thủy", "ninh thuỷ", "ninh phước", "ninh vân", "đông ninh hòa", "đông ninh hoà"], "exclude": KH_EXCLUDES},
    "Phường Hoà Thắng": {"keywords": ["ninh giang", "ninh hà", "ninh phú", "hòa thắng", "hoà thắng"], "exclude": KH_EXCLUDES},
    "Xã Bắc Ninh Hoà": {"keywords": ["ninh an", "ninh sơn", "ninh thọ", "bắc ninh hòa", "bắc ninh hoà"], "exclude": KH_EXCLUDES},
    "Xã Nam Ninh Hoà": {"keywords": ["ninh lộc", "ninh ích", "ninh hưng", "ninh tân", "nam ninh hòa", "nam ninh hoà"], "exclude": KH_EXCLUDES},
    "Bắc Nha Trang": {"keywords": ["lương sơn", "vĩnh lương", "văn đăng", "cát lợi", "võ tánh", "phạm văn đồng"], "exclude": ["ninh hòa", "ninh hoà", "vạn ninh", "cam ranh", "diên khánh", "cam lâm"]}
}

@st.cache_data
def find_true_header_and_cols(file_buffer):
    try:
        file_buffer.seek(0)
        df_temp = pd.read_excel(file_buffer, header=None, nrows=50)
        best_idx = 0; max_valid_cells = 0; std_to_original = {}
        for i, row in df_temp.iterrows():
            valid_cells = sum(1 for x in row.values if pd.notna(x) and str(x).strip() != '')
            if valid_cells > max_valid_cells:
                max_valid_cells = valid_cells; best_idx = i
                temp_map = {}
                for cell_val in row.values:
                    cell_str = str(cell_val).lower().strip(); orig_name = str(cell_val).strip()
                    if cell_str == 'nan' or not orig_name: continue
                    for std_name, alias_list in COLUMN_ALIASES.items():
                        if std_name not in temp_map and any(alias in cell_str for alias in alias_list):
                            temp_map[std_name] = orig_name; break
                std_to_original = temp_map
        file_buffer.seek(0)
        df_real = pd.read_excel(file_buffer, header=best_idx, nrows=5)
        cols = [c for c in df_real.columns.astype(str).str.strip().tolist() if not c.lower().startswith('unnamed') and c.lower() != 'nan' and c != 'None' and c != '']
        return cols, best_idx, std_to_original
    except Exception as e: return [], 0, {}

def get_new_address(old_address):
    if pd.isna(old_address) or str(old_address).strip() == 'nan': return None
    address_str = str(old_address).lower()
    for exc in GLOBAL_EXCLUDES:
        if exc in address_str: return None 
    for new_name, rule in mapping_rules.items():
        is_excluded = False
        for ex_word in rule["exclude"]:
            if ex_word in address_str: is_excluded = True; break
        if is_excluded: continue 
        for keyword in rule["keywords"]:
            if keyword in address_str: return new_name
    return None

def check_expired(date_str):
    if not date_str or date_str == '-': return False
    try:
        if '/' in date_str: return datetime.strptime(date_str, '%d/%m/%Y') < datetime.now()
        elif '-' in date_str: return datetime.strptime(date_str, '%Y-%m-%d') < datetime.now()
    except: return False

# =========================================================
# GIAO DIỆN CHUNG
# =========================================================
st.sidebar.markdown("## 🚢 QUẢN LÝ TÀU CÁ")
menu = st.sidebar.radio("MENU CHỨC NĂNG", ["🔍 Tra Cứu Thông Tin", "🔄 Đối Chiếu Dữ Liệu", "📊 Lọc & Thống Kê"])
st.sidebar.markdown("---")
st.sidebar.caption("Phát triển bởi Lê Hiếu - Trạm KN Ninh Hoà Vạn Ninh")

# =========================================================
# TRANG 1: TRA CỨU THÔNG TIN
# =========================================================
if menu == "🔍 Tra Cứu Thông Tin":
    st.markdown("### 🔍 TRA CỨU THÔNG TIN TÀU CÁ")
    
    col_db1, col_db2 = st.columns([3, 1])
    with col_db2:
        new_db = st.file_uploader("📥 Cập nhật CSDL Mới", type=['xlsx', 'xls'], help="Chỉ nạp khi có dữ liệu mới")
        if new_db:
            with open(DB_FILE, "wb") as f:
                f.write(new_db.getbuffer())
            st.success("✅ Cập nhật CSDL thành công!")
    
    with col_db1:
        if os.path.exists(DB_FILE):
            mtime = os.path.getmtime(DB_FILE)
            dt_mtime = datetime.fromtimestamp(mtime).strftime('%H:%M %d/%m/%Y')
            st.info(f"✅ Hệ thống đang sử dụng CSDL được cập nhật lúc: **{dt_mtime}**")
            
            with open(DB_FILE, "rb") as f:
                db_bytes = io.BytesIO(f.read())
            db_cols, db_hdr_idx, db_std_map = find_true_header_and_cols(db_bytes)
            
            # Form tìm kiếm
            col_s1, col_s2 = st.columns([1, 3])
            search_type = col_s1.selectbox("Tìm theo", ["Số đăng ký", "Tên chủ tàu", "Tất cả"])
            keyword = col_s2.text_input("Nhập từ khóa tìm kiếm (Gõ và Enter):")
            
            if keyword:
                db_bytes.seek(0)
                df_db = pd.read_excel(db_bytes, header=db_hdr_idx)
                df_db.columns = df_db.columns.astype(str).str.strip()
                
                col_dk = db_std_map.get('SO_DANG_KY')
                col_ten = db_std_map.get('CHU_TAU')
                
                keyword = keyword.strip().lower()
                if search_type == "Số đăng ký" and col_dk and col_dk in df_db.columns:
                    res = df_db[df_db[col_dk].astype(str).str.lower().str.contains(keyword, na=False, regex=False)]
                elif search_type == "Tên chủ tàu" and col_ten and col_ten in df_db.columns:
                    res = df_db[df_db[col_ten].astype(str).str.lower().str.contains(keyword, na=False, regex=False)]
                else:
                    mask = df_db.apply(lambda row: row.astype(str).str.lower().str.contains(keyword, na=False, regex=False).any(), axis=1)
                    res = df_db[mask]
                
                if len(res) == 0:
                    st.warning(f"Không tìm thấy kết quả nào cho từ khóa: '{keyword}'")
                else:
                    # Chuẩn hóa Dataframe để hiển thị
                    T3_COL_MAP = {
                        "Số đăng ký": "SO_DANG_KY", "Tên chủ tàu": "CHU_TAU", "SĐT": "SDT", "CCCD": "CCCD",
                        "Địa phương": "XA_PHUONG", "Địa chỉ": "DIA_CHI", "Lmax": "LMAX", "Công suất": "CONG_SUAT", 
                        "Hạn Đăng kiểm": "HAN_DK", "Hạn GPKTTS": "HAN_GP"
                    }
                    display_list = []
                    for idx, row in res.iterrows():
                        row_data = {}
                        for col_title, c_alias in T3_COL_MAP.items():
                            orig_col = db_std_map.get(c_alias)
                            if not orig_col and c_alias in ['HAN_GP', 'HAN_DK']: orig_col = db_std_map.get('NGAY_HET_HAN')
                            
                            val = None
                            if orig_col and orig_col in df_db.columns: val = row[orig_col]
                            elif c_alias == 'XA_PHUONG':
                                dc_col = db_std_map.get('DIA_CHI')
                                if dc_col and dc_col in df_db.columns: val = get_new_address(row[dc_col])

                            if pd.notna(val) and not isinstance(val, (datetime, pd.Timestamp)) and c_alias not in ['SDT', 'CCCD', 'HAN_GP', 'HAN_DK']:
                                if str(val).endswith('.0'): val = str(val)[:-2]
                            if c_alias == 'CCCD' and pd.notna(val):
                                v_str = str(val).strip().split('.')[0]
                                if v_str.lower() not in ['nan', 'none', ''] and v_str.isdigit(): val = v_str.zfill(12)
                            elif c_alias == 'SDT' and pd.notna(val):
                                v_str = str(val).strip().split('.')[0]
                                if v_str.lower() not in ['nan', 'none', ''] and v_str.isdigit():
                                    if len(v_str) >= 9 and v_str[0] != '0': val = '0' + v_str
                            elif c_alias in ['HAN_GP', 'HAN_DK'] and pd.notna(val):
                                try:
                                    parsed = pd.to_datetime(val, errors='coerce', dayfirst=True)
                                    if pd.notna(parsed): val = parsed.strftime('%d/%m/%Y')
                                    else: val = str(val).split(' ')[0].strip()
                                except: pass
                            row_data[col_title] = str(val) if pd.notna(val) and str(val).lower() not in ['nan', 'none'] else "-"
                        display_list.append(row_data)
                    
                    df_display = pd.DataFrame(display_list)
                    
                    st.markdown("---")
                    st.markdown("### KẾT QUẢ TÌM KIẾM")
                    
                    # Tách làm 2 cột: Bảng và Thẻ
                    col_table, col_card = st.columns([1.3, 1])
                    
                    with col_table:
                        # Sử dụng dataframe với on_select nếu Streamlit hỗ trợ, nếu không thì dùng Selectbox
                        st.dataframe(df_display, use_container_width=True, hide_index=True)
                        
                    with col_card:
                        selected_vessel = st.selectbox("📌 Chọn Số đăng ký để xem Thẻ thông tin chi tiết:", df_display['Số đăng ký'])
                        
                        if selected_vessel:
                            v_data = df_display[df_display['Số đăng ký'] == selected_vessel].iloc[0]
                            
                            # Xử lý màu sắc hạn
                            c_hdk = "#dc3545" if check_expired(v_data['Hạn Đăng kiểm']) else "#198754"
                            c_hgp = "#dc3545" if check_expired(v_data['Hạn GPKTTS']) else "#198754"
                            
                            # HTML/CSS Card
                            card_html = f"""
                            <div style="border: 1px solid #dee2e6; border-radius: 8px; background: white; box-shadow: 0 4px 6px rgba(0,0,0,0.1); margin-bottom: 20px;">
                                <div style="background: #0d6efd; color: white; padding: 20px; border-radius: 8px 8px 0 0;">
                                    <div style="font-weight: bold; font-size: 13px; text-transform: uppercase;">CHI TIẾT TÀU CÁ</div>
                                    <div style="font-weight: bold; font-size: 26px; margin: 5px 0;">{v_data['Số đăng ký']}</div>
                                    <div style="background: white; color: #1864ab; display: inline-block; padding: 4px 12px; border-radius: 12px; font-size: 13px; font-weight: bold;">
                                        📍 {v_data['Địa phương']}
                                    </div>
                                </div>
                                
                                <div style="padding: 20px; font-family: sans-serif;">
                                    <div style="display: flex; margin-bottom: 15px;">
                                        <div style="flex: 1;">
                                            <div style="color: #6c757d; font-size: 12px; font-weight: bold;">👤 CHỦ PHƯƠNG TIỆN</div>
                                            <div style="font-size: 15px; font-weight: bold; color: #212529;">{v_data['Tên chủ tàu']}</div>
                                        </div>
                                        <div style="flex: 1;">
                                            <div style="color: #6c757d; font-size: 12px; font-weight: bold;">💳 SỐ CCCD/CMND</div>
                                            <div style="font-size: 15px; font-weight: bold; color: #212529;">{v_data['CCCD']}</div>
                                        </div>
                                    </div>
                                    
                                    <div style="display: flex; margin-bottom: 15px;">
                                        <div style="flex: 1;">
                                            <div style="color: #6c757d; font-size: 12px; font-weight: bold;">📞 SỐ ĐIỆN THOẠI</div>
                                            <div style="font-size: 15px; font-weight: bold; color: #212529;">{v_data['SĐT']}</div>
                                        </div>
                                        <div style="flex: 1;">
                                            <div style="color: #6c757d; font-size: 12px; font-weight: bold;">📏 CHIỀU DÀI LMAX</div>
                                            <div style="font-size: 15px; font-weight: bold; color: #212529;">{v_data['Lmax']} m</div>
                                        </div>
                                    </div>
                                    
                                    <div style="margin-bottom: 20px;">
                                        <div style="color: #6c757d; font-size: 12px; font-weight: bold;">⚡ CÔNG SUẤT (KW)</div>
                                        <div style="font-size: 15px; font-weight: bold; color: #212529;">{v_data['Công suất']} KW</div>
                                    </div>
                                    
                                    <div style="margin-bottom: 25px;">
                                        <div style="color: #6c757d; font-size: 12px; font-weight: bold;">🏠 ĐỊA CHỈ THƯỜNG TRÚ</div>
                                        <div style="font-size: 14px; font-weight: bold; color: #212529;">{v_data['Địa chỉ']}</div>
                                    </div>
                                    
                                    <div style="background: #f8f9fa; border: 1px solid #dee2e6; border-radius: 6px; padding: 12px 15px; margin-bottom: 10px; display: flex; justify-content: space-between; align-items: center;">
                                        <div style="color: #495057; font-size: 13px; font-weight: bold;">🗓 HẠN GIẤY PHÉP KTTS</div>
                                        <div style="font-size: 14px; font-weight: bold; color: {c_hgp};">{v_data['Hạn GPKTTS']}</div>
                                    </div>
                                    
                                    <div style="background: #f8f9fa; border: 1px solid #dee2e6; border-radius: 6px; padding: 12px 15px; display: flex; justify-content: space-between; align-items: center;">
                                        <div style="color: #495057; font-size: 13px; font-weight: bold;">🗓 HẠN ĐĂNG KIỂM</div>
                                        <div style="font-size: 14px; font-weight: bold; color: {c_hdk};">{v_data['Hạn Đăng kiểm']}</div>
                                    </div>
                                </div>
                            </div>
                            """
                            st.markdown(card_html, unsafe_allow_html=True)
        else:
            st.error("⚠️ Hệ thống chưa có dữ liệu. Vui lòng tải lên file CSDL ở khung bên trên!")


# =========================================================
# TRANG 2: ĐỐI CHIẾU DỮ LIỆU
# =========================================================
elif menu == "🔄 Đối Chiếu Dữ Liệu":
    st.markdown("### 🔄 ĐỐI CHIẾU VÀ ĐẮP THÊM DỮ LIỆU")
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("##### 1. File Gốc (Cần đắp thêm)")
        file_src = st.file_uploader("Chọn File Excel Gốc", type=['xlsx', 'xls'], key="src")
    with col2:
        st.markdown("##### 2. File Đích (Nguồn lấy dữ liệu)")
        file_tgts = st.file_uploader("Chọn (các) File Đích", type=['xlsx', 'xls'], accept_multiple_files=True, key="tgt")
        
    if file_src and file_tgts:
        st.markdown("---")
        st.markdown("##### 3. Thiết lập Đối chiếu")
        
        # Đọc Header Gốc
        cols_src, hdr_src, _ = find_true_header_and_cols(file_src)
        
        # Đọc gộp Header Đích
        all_tgt_cols = []
        for f in file_tgts:
            c_t, _, _ = find_true_header_and_cols(f)
            for c in c_t:
                if c not in all_tgt_cols: all_tgt_cols.append(c)
                
        col_k1, col_k2 = st.columns(2)
        
        default_k_src = 0
        for i, c in enumerate(cols_src):
            if any(a in c.lower() for a in COLUMN_ALIASES['SO_DANG_KY']): default_k_src = i; break
            
        k_src = col_k1.selectbox("Cột Nối Gốc:", cols_src, index=default_k_src)
        k_tgt = col_k2.selectbox("Cột Nối Đích:", ["<Tự động nhận diện bằng AI>"] + all_tgt_cols)
        
        vals_to_get = st.multiselect("Chọn (các) Cột ở File Đích muốn lấy sang File Gốc:", all_tgt_cols)
        
        if vals_to_get:
            if st.button("🚀 CHẠY ĐỐI CHIẾU NGAY", type="primary"):
                with st.spinner("Đang xử lý dữ liệu..."):
                    try:
                        file_src.seek(0)
                        df_src = pd.read_excel(file_src, header=hdr_src)
                        df_src.columns = df_src.columns.astype(str).str.strip()
                        df_src['_key_match'] = df_src[k_src].astype(str).str.strip().str.upper()
                        
                        df_tgt_list = []
                        for p in file_tgts:
                            p.seek(0)
                            cols, hdr_idx, std_to_original = find_true_header_and_cols(p, is_scan=True)
                            local_key = None
                            if k_tgt == "<Tự động nhận diện bằng AI>":
                                for std, orig in std_to_original.items():
                                    if std == 'SO_DANG_KY': local_key = orig; break
                            else:
                                local_key = k_tgt if k_tgt in cols else None

                            if not local_key: continue
                            
                            local_vals = [c for c in vals_to_get if c in cols]
                            if not local_vals: continue
                            
                            p.seek(0)
                            df_t = pd.read_excel(p, header=hdr_idx)
                            df_t.columns = df_t.columns.astype(str).str.strip()
                            df_t['_key_match'] = df_t[local_key].astype(str).str.strip().str.upper()
                            df_t_unique = df_t.drop_duplicates(subset=['_key_match'])
                            
                            short_fname = p.name[:15]
                            rename_dict = {c: f"{c} ({short_fname}...)" for c in local_vals}
                            df_t_sub = df_t_unique[['_key_match'] + local_vals].rename(columns=rename_dict)
                            df_src = pd.merge(df_src, df_t_sub, on='_key_match', how='left')
                            
                        df_src.drop(columns=['_key_match'], inplace=True)
                        
                        # Xuất file
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                            df_src.to_excel(writer, index=False, sheet_name='Ket_qua')
                        output.seek(0)
                        
                        now_str = datetime.now().strftime("%d%m%Y_%Hh%Mm")
                        st.success("✅ Đối chiếu thành công! Bấm nút bên dưới để lưu file.")
                        st.download_button(
                            label="📥 Tải File Kết Quả",
                            data=output,
                            file_name=f"Ket_qua_Doi_chieu_{now_str}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    except Exception as e:
                        st.error(f"❌ Lỗi xử lý: {e}")


# =========================================================
# TRANG 3: LỌC BÁO CÁO & THỐNG KÊ
# =========================================================
elif menu == "📊 Lọc & Thống Kê":
    st.markdown("### 📊 LỌC DỮ LIỆU & THỐNG KÊ TÀU CÁ")
    
    file_t1 = st.file_uploader("1. Chọn Dữ liệu đầu vào (File Excel)", type=['xlsx', 'xls'])
    if file_t1:
        cols_t1, hdr_t1, std_to_original_t1 = find_true_header_and_cols(file_t1, is_scan=True)
        if cols_t1:
            st.markdown("##### 2. Tùy chọn Cột Báo Cáo")
            default_selections = [v for v in std_to_original_t1.values() if v in cols_t1]
            selected_cols = st.multiselect("Chọn các cột muốn xuất (Kéo thả để sắp xếp thứ tự):", ["Xã/Phường mới"] + cols_t1, default=default_selections + ["Xã/Phường mới"])
            
            st.markdown("##### 3. Điều Kiện Lọc")
            col_f1, col_f2, col_f3 = st.columns(3)
            
            with col_f1:
                selected_commune = st.selectbox("Địa phương:", ["Tất cả", "Xã Đại Lãnh", "Xã Tu Bông", "Xã Vạn Hưng", "Xã Vạn Ninh", "Xã Vạn Thắng", "Phường Đông Ninh Hoà", "Phường Hoà Thắng", "Xã Bắc Ninh Hoà", "Xã Nam Ninh Hoà", "Bắc Nha Trang"])
            
            with col_f2:
                date_default = 0
                if 'HAN_DK' in std_to_original_t1 and std_to_original_t1['HAN_DK'] in cols_t1: date_default = cols_t1.index(std_to_original_t1['HAN_DK'])
                elif 'HAN_GP' in std_to_original_t1 and std_to_original_t1['HAN_GP'] in cols_t1: date_default = cols_t1.index(std_to_original_t1['HAN_GP'])
                selected_date_col = st.selectbox("Cột Mốc Tính Hạn:", cols_t1, index=date_default)
            
            with col_f3:
                selected_date = st.text_input("Lọc đến ngày (Ví dụ: 30/06/2026. Bỏ trống để lấy tất cả):")
            
            split_expired = st.checkbox("Tách mỗi Xã/Phường thành 1 Sheet (Đối với Tàu Hết hạn)", value=False)
            
            if st.button("🚀 BẮT ĐẦU XỬ LÝ & TẠO BÁO CÁO", type="primary"):
                if not selected_cols:
                    st.warning("Vui lòng chọn ít nhất 1 cột báo cáo!")
                else:
                    with st.spinner("Đang phân tích dữ liệu bằng AI..."):
                        try:
                            file_t1.seek(0)
                            df_raw = pd.read_excel(file_t1, header=hdr_t1)
                            df_raw.columns = df_raw.columns.astype(str).str.strip()
                            
                            col_diachi_goc = std_to_original_t1.get('DIA_CHI', '')
                            if col_diachi_goc: df_raw['Xã/Phường mới'] = df_raw[col_diachi_goc].apply(get_new_address)
                            else: st.error("❌ LỖI: Không tìm thấy Cột Địa chỉ trong file gốc!"); st.stop()
                                
                            df_filtered = df_raw[df_raw['Xã/Phường mới'].notna()].copy()
                            if selected_commune != "Tất cả": df_filtered = df_filtered[df_filtered['Xã/Phường mới'] == selected_commune]

                            col_cccd = std_to_original_t1.get('CCCD', '')
                            if col_cccd and col_cccd in df_filtered.columns:
                                def format_cccd_export(x):
                                    if pd.isna(x): return x
                                    v_str = str(x).strip().split('.')[0]
                                    if v_str.lower() in ['nan', 'none', ''] or not v_str.isdigit(): return x
                                    return v_str.zfill(12) 
                                df_filtered[col_cccd] = df_filtered[col_cccd].apply(format_cccd_export)

                            if selected_date_col in df_filtered.columns:
                                try: parsed_dates = pd.to_datetime(df_filtered[selected_date_col], dayfirst=True, format='mixed', errors='coerce')
                                except ValueError: parsed_dates = pd.to_datetime(df_filtered[selected_date_col], dayfirst=True, errors='coerce')
                                mask_old = (parsed_dates.dt.year < 1950) & parsed_dates.notna()
                                if mask_old.any(): parsed_dates.loc[mask_old] = parsed_dates.loc[mask_old].apply(lambda x: x.replace(year=x.year + 100))
                                df_filtered['Ngày_dt_temp'] = parsed_dates
                                if selected_date.strip() != "":
                                    target_date = pd.to_datetime(selected_date.strip(), dayfirst=True)
                                    df_filtered = df_filtered[df_filtered['Ngày_dt_temp'] <= target_date]

                            if len(df_filtered) == 0: 
                                st.warning("⚠️ Không có tàu nào thỏa mãn điều kiện!"); st.stop()

                            final_dict = {col: df_filtered[col] for col in selected_cols if col in df_filtered.columns}
                            df_final = pd.DataFrame(final_dict); df_final.insert(0, 'TT', range(1, len(df_final) + 1))

                            current_date = pd.Timestamp.now().normalize()
                            if selected_date_col in df_filtered.columns: df_filtered['_da_het_han'] = df_filtered['Ngày_dt_temp'] < current_date
                            else: df_filtered['_da_het_han'] = False

                            if 'LMAX' in std_to_original_t1: df_filtered['Lmax_num'] = pd.to_numeric(df_filtered[std_to_original_t1['LMAX']], errors='coerce')
                            else: df_filtered['Lmax_num'] = None

                            def phan_loai_lmax(lmax):
                                if pd.isna(lmax): return 'Không rõ'
                                if lmax < 6: return '<6'
                                elif 6 <= lmax < 12: return '6 đến <12'
                                elif 12 <= lmax < 15: return '12 đến <15'
                                elif 15 <= lmax < 24: return '15 đến <24'
                                else: return '>=24'

                            df_filtered['Nhom_Lmax'] = df_filtered['Lmax_num'].apply(phan_loai_lmax)
                            col_id = std_to_original_t1.get('SO_DANG_KY', df_filtered.columns[0])
                            df_thong_ke_main = df_filtered.groupby('Xã/Phường mới').agg(Tong_tau=(col_id, 'count'), Tau_het_han=('_da_het_han', 'sum')).reset_index()
                            lmax_pivot = pd.crosstab(df_filtered['Xã/Phường mới'], df_filtered['Nhom_Lmax']).reset_index()
                            df_thong_ke = pd.merge(df_thong_ke_main, lmax_pivot, on='Xã/Phường mới', how='left')

                            for c in ['<6', '6 đến <12', '12 đến <15', '15 đến <24', '>=24', 'Không rõ']:
                                if c not in df_thong_ke.columns: df_thong_ke[c] = 0

                            df_thong_ke.rename(columns={'Tong_tau': 'Tổng số tàu', 'Tau_het_han': 'Tàu hết hạn'}, inplace=True)
                            final_cols_tk = ['Xã/Phường mới', 'Tổng số tàu', 'Tàu hết hạn']
                            if df_thong_ke['<6'].sum() > 0: final_cols_tk.append('<6')
                            final_cols_tk.extend(['6 đến <12', '12 đến <15', '15 đến <24', '>=24'])
                            if df_thong_ke['Không rõ'].sum() > 0: final_cols_tk.append('Không rõ')
                            df_thong_ke = df_thong_ke[final_cols_tk]

                            tong_cong_row = {'Xã/Phường mới': 'TỔNG CỘNG'}
                            for c in df_thong_ke.columns:
                                if c != 'Xã/Phường mới': tong_cong_row[c] = df_thong_ke[c].sum()
                            df_thong_ke = pd.concat([df_thong_ke, pd.DataFrame([tong_cong_row])], ignore_index=True)

                            df_het_han = df_filtered[df_filtered['_da_het_han'] == True].copy()
                            cols_to_export_hh = [c for c in selected_cols if c in df_het_han.columns]
                            if selected_date_col not in cols_to_export_hh and selected_date_col in df_het_han.columns: cols_to_export_hh.append(selected_date_col)
                            df_het_han_export = df_het_han[cols_to_export_hh].copy()
                            if not df_het_han_export.empty: df_het_han_export.insert(0, 'TT', range(1, len(df_het_han_export) + 1))

                            # Format và xuất BytesIO
                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                df_final.to_excel(writer, sheet_name='Danh sách chi tiết', index=False)
                                df_thong_ke.to_excel(writer, sheet_name='Bảng thống kê', index=False)
                                
                                if not df_het_han_export.empty: 
                                    df_het_han_export.to_excel(writer, sheet_name='Tàu Hết Hạn (Tổng)', index=False)
                                    if split_expired:
                                        unique_locs = df_het_han['Xã/Phường mới'].dropna().unique()
                                        for loc in unique_locs:
                                            df_loc = df_het_han[df_het_han['Xã/Phường mới'] == loc].copy()
                                            df_loc_export = df_loc[cols_to_export_hh].copy()
                                            df_loc_export.insert(0, 'TT', range(1, len(df_loc_export) + 1))
                                            sn = f"HH_{str(loc).replace('/', '_').replace(chr(92), '_')}"[:31] 
                                            df_loc_export.to_excel(writer, sheet_name=sn, index=False)
                                
                                workbook = writer.book
                                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                                center_align = Alignment(horizontal="center", vertical="center")
                                left_align = Alignment(horizontal="left", vertical="center")
                                header_font = Font(color="FFFFFF", bold=True)
                                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                                ws_chi_tiet = writer.sheets['Danh sách chi tiết']
                                for col_num, cell in enumerate(ws_chi_tiet[1], 1):
                                    cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align; cell.border = thin_border
                                    ws_chi_tiet.column_dimensions[cell.column_letter].width = 20
                                ws_chi_tiet.column_dimensions['A'].width = 8 

                                ws_thong_ke = writer.sheets['Bảng thống kê']
                                total_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
                                total_font = Font(bold=True, color="C0504D")
                                for col_num, cell in enumerate(ws_thong_ke[1], 1):
                                    cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align; cell.border = thin_border
                                    ws_thong_ke.column_dimensions[cell.column_letter].width = 30 if col_num == 1 else 15

                                for row in range(2, ws_thong_ke.max_row + 1):
                                    for c in range(1, ws_thong_ke.max_column + 1):
                                        cell = ws_thong_ke.cell(row=row, column=c)
                                        cell.border = thin_border
                                        cell.alignment = left_align if c == 1 else center_align
                                        if row == ws_thong_ke.max_row: cell.fill = total_fill; cell.font = total_font
                                            
                                if not df_het_han_export.empty:
                                    danger_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
                                    for sheet_name in writer.sheets:
                                        if sheet_name.startswith('Tàu Hết Hạn') or sheet_name.startswith('HH_'):
                                            ws_hh = writer.sheets[sheet_name]
                                            for col_num, cell in enumerate(ws_hh[1], 1):
                                                cell.fill = danger_fill; cell.font = header_font; cell.alignment = center_align; cell.border = thin_border
                                                ws_hh.column_dimensions[cell.column_letter].width = 20
                                            ws_hh.column_dimensions['A'].width = 8 
                            
                            output.seek(0)
                            st.success(f"🎉 Đã xử lý xong {tong_cong_row['Tổng số tàu']} tàu!")
                            
                            now_str = datetime.now().strftime("%d%m%Y_%Hh%Mm")
                            base_name = "DS_tau_ca" if selected_commune == "Tất cả" else f"DS_tau_ca_{selected_commune.replace(' ', '_')}"
                            if selected_date != "": base_name += f"_Han_{selected_date.replace('/', '')}"
                            
                            st.download_button(
                                label="📥 Tải File Báo Cáo (.xlsx)",
                                data=output,
                                file_name=f"{base_name}_NHVN_{now_str}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        except Exception as e:
                            st.error(f"❌ Lỗi xử lý: {e}")
